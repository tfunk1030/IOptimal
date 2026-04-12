"""Parse individual aero map xlsx files into structured numpy arrays.

Two xlsx formats exist across the 5 GTP cars:

Labeled (BMW, Cadillac, Acura, Porsche):
  - Sheet "Raw Data" / "Data tables" / "AeroBalance"
  - Row 1: rear ride height values (columns B onward)
  - Col A: front ride height values (rows 2 onward)
  - Left block: DF balance (%)
  - Right block (after 3-4 empty cols): L/D ratio
  - Right block has its own front RH labels in a separate column

Unlabeled (Ferrari):
  - Sheet "AeroBalance"
  - No headers — row 1 is empty
  - Data starts at row 2, col 2
  - Same 51-row × 46-col grid, just without labels
  - Front RH assumed 25–75mm (1mm steps), rear RH assumed 5–50mm (1mm steps)
"""

import json
import logging
from pathlib import Path

import numpy as np
import openpyxl

logger = logging.getLogger(__name__)


# Map from car directory name to the data sheet name
_SHEET_NAMES = {
    "BMWGTPAeroMap": "Raw Data",
    "CadallliacGTPAeromap": "Data tables",
    "Acuragtpaeromap": "AeroBalance",
    "Ferrarigtpaeromap": "AeroBalance",
    "Porschegtpaeromap": "AeroBalance",
}

# Map from directory name to canonical car name
CAR_NAMES = {
    "BMWGTPAeroMap": "bmw",
    "CadallliacGTPAeromap": "cadillac",
    "Acuragtpaeromap": "acura",
    "Ferrarigtpaeromap": "ferrari",
    "Porschegtpaeromap": "porsche",
}


def _find_data_sheet(wb: openpyxl.Workbook, car_dir: str) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the worksheet containing the raw aero data."""
    target = _SHEET_NAMES.get(car_dir)
    if target and target in wb.sheetnames:
        return wb[target]
    # Fallback: try common names
    for name in ("Raw Data", "Data tables", "Data Table", "AeroBalance"):
        if name in wb.sheetnames:
            sheet = wb[name]
            # Skip chart sheets
            if hasattr(sheet, "max_row"):
                return sheet
    raise ValueError(f"No data sheet found in workbook. Sheets: {wb.sheetnames}")


def _align_ld_to_balance(
    ld: np.ndarray,
    rear_rh_ld: list[float],
    rear_rh_balance: np.ndarray,
) -> tuple[np.ndarray, list[float]]:
    """Pad L/D grid to match balance rear_rh axis when columns are missing.

    Some xlsx files (e.g., Porsche wing 13) have a front-RH label column with
    a spurious numeric header that eats one rear_rh slot.  When the L/D block
    is missing leading columns that the balance block has, extrapolate them
    linearly from the first two available L/D columns.
    """
    ld_rh_to_idx = {rh: i for i, rh in enumerate(rear_rh_ld)}
    bal_rh = list(rear_rh_balance)
    if all(rh in ld_rh_to_idx for rh in bal_rh):
        return ld, rear_rh_ld

    n_front = ld.shape[0]
    new_ld = np.zeros((n_front, len(bal_rh)))
    ld_rh_arr = np.array(rear_rh_ld)

    for j, rh in enumerate(bal_rh):
        if rh in ld_rh_to_idx:
            new_ld[:, j] = ld[:, ld_rh_to_idx[rh]]
        else:
            # Linear extrapolation from the two nearest available columns
            if rh < ld_rh_arr[0]:
                # Extrapolate left from columns 0 and 1
                slope = (ld[:, 1] - ld[:, 0]) / (ld_rh_arr[1] - ld_rh_arr[0])
                new_ld[:, j] = ld[:, 0] + slope * (rh - ld_rh_arr[0])
            else:
                # Extrapolate right from last two columns
                slope = (ld[:, -1] - ld[:, -2]) / (ld_rh_arr[-1] - ld_rh_arr[-2])
                new_ld[:, j] = ld[:, -1] + slope * (rh - ld_rh_arr[-1])

    return new_ld, bal_rh


def _detect_format(ws) -> str:
    """Detect whether the sheet has labeled headers or is unlabeled (Ferrari)."""
    # Check if row 1 has numeric values in columns B+ (labeled format)
    for c in range(2, 10):
        v = ws.cell(row=1, column=c).value
        if v is not None and isinstance(v, (int, float)):
            return "labeled"
    return "unlabeled"


def _parse_labeled(ws) -> dict:
    """Parse a labeled-format sheet (BMW, Cadillac, Acura, Porsche)."""
    # Row 1, cols B onward: rear ride heights for balance block
    rear_rh_balance = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            break
        rear_rh_balance.append(float(v))

    n_rear = len(rear_rh_balance)

    # Col A, rows 2 onward: front ride heights
    front_rh = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            break
        front_rh.append(float(v))

    n_front = len(front_rh)

    # Balance data block: rows 2..(1+n_front), cols 2..(1+n_rear)
    balance = np.zeros((n_front, n_rear))
    for i in range(n_front):
        for j in range(n_rear):
            v = ws.cell(row=2 + i, column=2 + j).value
            balance[i, j] = float(v) if v is not None else np.nan

    # Find L/D block: scan rightward from end of balance block.
    # Must find a column where BOTH row 1 (header) and row 2 (data) have numeric values
    # AND row 2 actually looks like an L/D value (< 10), not a front-RH label (25-75).
    # Some xlsx files (Porsche wing 13) have the front-RH label column with a numeric
    # value in row 1 (e.g., "5") which would otherwise fool the detection.
    ld_start_col = None
    for c in range(2 + n_rear, ws.max_column + 1):
        v1 = ws.cell(row=1, column=c).value
        v2 = ws.cell(row=2, column=c).value
        if (v1 is not None and isinstance(v1, (int, float))
                and v2 is not None and isinstance(v2, (int, float))
                and float(v2) < 10.0):
            ld_start_col = c
            break

    if ld_start_col is None:
        raise ValueError("Could not find L/D data block")

    # Read rear RH headers for L/D block (row 1, starting at ld_start_col)
    rear_rh_ld = []
    for c in range(ld_start_col, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            break
        rear_rh_ld.append(float(v))

    # L/D data
    ld = np.zeros((n_front, len(rear_rh_ld)))
    for i in range(n_front):
        for j in range(len(rear_rh_ld)):
            v = ws.cell(row=2 + i, column=ld_start_col + j).value
            ld[i, j] = float(v) if v is not None else np.nan

    rear_rh_bal = np.array(rear_rh_balance)

    # If the L/D block has fewer rear_rh columns than balance (e.g., Porsche wing 13
    # is missing rrh=5mm), pad by linear extrapolation from the first two L/D columns.
    if len(rear_rh_ld) < n_rear:
        ld, rear_rh_ld = _align_ld_to_balance(ld, rear_rh_ld, rear_rh_bal)

    return {
        "front_rh": np.array(front_rh),
        "rear_rh": rear_rh_bal,
        "balance": balance,
        "ld": ld,
    }


def _parse_unlabeled(ws) -> dict:
    """Parse an unlabeled-format sheet (Ferrari).

    No header row or label column. Data starts at row 2, col 2.
    Grid is 51 rows × 46 cols for each block.
    Assumed: front RH 25–75mm, rear RH 5–50mm.
    """
    # Count contiguous data columns in row 2 starting from col 2
    n_rear = 0
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is None:
            break
        n_rear += 1

    # Count contiguous data rows starting from row 2
    n_front = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v is None:
            break
        n_front += 1

    # Infer ride height axes (standard GTP grid)
    front_rh = np.arange(25, 25 + n_front, dtype=float)  # 25, 26, ..., 75
    rear_rh = np.arange(5, 5 + n_rear, dtype=float)  # 5, 6, ..., 50

    # Balance block
    balance = np.zeros((n_front, n_rear))
    for i in range(n_front):
        for j in range(n_rear):
            v = ws.cell(row=2 + i, column=2 + j).value
            balance[i, j] = float(v) if v is not None else np.nan

    # Find L/D block: scan rightward past the gap
    ld_start_col = None
    for c in range(2 + n_rear, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is not None and isinstance(v, (int, float)):
            ld_start_col = c
            break

    if ld_start_col is None:
        raise ValueError("Could not find L/D data block in unlabeled sheet")

    # Count L/D columns
    n_rear_ld = 0
    for c in range(ld_start_col, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is None:
            break
        n_rear_ld += 1

    ld = np.zeros((n_front, n_rear_ld))
    for i in range(n_front):
        for j in range(n_rear_ld):
            v = ws.cell(row=2 + i, column=ld_start_col + j).value
            ld[i, j] = float(v) if v is not None else np.nan

    return {
        "front_rh": front_rh,
        "rear_rh": rear_rh,
        "balance": balance,
        "ld": ld,
    }


def parse_aero_xlsx(filepath: str | Path) -> dict:
    """Parse a single aero map xlsx file.

    Returns dict with:
        car: str - canonical car name
        wing_angle: float - wing angle from filename
        front_rh: np.ndarray - front ride height axis (mm)
        rear_rh: np.ndarray - rear ride height axis (mm)
        balance: np.ndarray - DF balance (%) grid [front_rh × rear_rh]
        ld: np.ndarray - L/D ratio grid [front_rh × rear_rh]
    """
    filepath = Path(filepath)
    car_dir = filepath.parent.name
    car_name = CAR_NAMES.get(car_dir, car_dir)

    # Extract wing angle from filename
    wing_angle = _extract_wing_angle(filepath.name)

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=False)
    ws = _find_data_sheet(wb, car_dir)

    fmt = _detect_format(ws)
    if fmt == "labeled":
        data = _parse_labeled(ws)
    else:
        data = _parse_unlabeled(ws)

    wb.close()

    _validate_parsed_data(data, car_name, wing_angle, filepath)

    return {
        "car": car_name,
        "wing_angle": wing_angle,
        **data,
    }


def _validate_parsed_data(
    data: dict, car: str, wing: float, filepath: Path,
) -> None:
    """Validate parsed aero data has physically plausible values."""
    ld = data["ld"]
    balance = data["balance"]

    ld_min, ld_max = float(np.nanmin(ld)), float(np.nanmax(ld))
    if ld_min < 1.0 or ld_max > 6.0:
        raise ValueError(
            f"Aero map {car} wing {wing} ({filepath.name}): "
            f"L/D values out of range [1.0, 6.0]: "
            f"min={ld_min:.3f}, max={ld_max:.3f}"
        )

    bal_min, bal_max = float(np.nanmin(balance)), float(np.nanmax(balance))
    if bal_min < 10.0 or bal_max > 90.0:
        raise ValueError(
            f"Aero map {car} wing {wing} ({filepath.name}): "
            f"DF balance values out of range [10.0, 90.0]: "
            f"min={bal_min:.2f}, max={bal_max:.2f}"
        )

    if ld.shape != balance.shape:
        raise ValueError(
            f"Aero map {car} wing {wing} ({filepath.name}): "
            f"L/D shape {ld.shape} != balance shape {balance.shape}"
        )


def _extract_wing_angle(filename: str) -> float:
    """Extract wing angle from filename like 'Aero data 17 wing BMW LMDH.xlsx'."""
    # Pattern: number before "wing" in the filename
    parts = filename.lower().replace(".xlsx", "").split()
    for i, part in enumerate(parts):
        if part == "wing" and i > 0:
            try:
                return float(parts[i - 1])
            except ValueError:
                pass
    raise ValueError(f"Could not extract wing angle from filename: {filename}")


# ---------------------------------------------------------------------------
# Batch parsing: xlsx → npz + json
# ---------------------------------------------------------------------------

_AERO_SRC_DIR = Path(__file__).parent.parent / "data" / "aeromaps"
_AERO_OUT_DIR = Path(__file__).parent.parent / "data" / "aeromaps_parsed"


def parse_all_cars() -> None:
    """Parse all car aero map xlsx files and write npz + json to data/aeromaps_parsed/.

    Usage::

        python -m aero_model.parse_xlsx
    """
    _AERO_OUT_DIR.mkdir(parents=True, exist_ok=True)

    for car_dir_path in sorted(_AERO_SRC_DIR.iterdir()):
        if not car_dir_path.is_dir():
            continue
        car_name = CAR_NAMES.get(car_dir_path.name)
        if car_name is None:
            continue

        xlsx_files = sorted(car_dir_path.glob("*.xlsx"))
        if not xlsx_files:
            continue

        arrays: dict[str, np.ndarray] = {}
        wing_angles: list[float] = []
        shared_front_rh = None
        shared_rear_rh = None

        for xlsx in xlsx_files:
            parsed = parse_aero_xlsx(xlsx)
            wing = parsed["wing_angle"]
            wing_angles.append(wing)
            arrays[f"balance_{wing}"] = parsed["balance"]
            arrays[f"ld_{wing}"] = parsed["ld"]

            if shared_front_rh is None:
                shared_front_rh = parsed["front_rh"]
                shared_rear_rh = parsed["rear_rh"]

        arrays["front_rh"] = shared_front_rh
        arrays["rear_rh"] = shared_rear_rh

        npz_path = _AERO_OUT_DIR / f"{car_name}_aero.npz"
        np.savez_compressed(str(npz_path), **arrays)

        meta = {
            "car": car_name,
            "wing_angles": sorted(wing_angles),
            "front_rh_range": [float(shared_front_rh[0]), float(shared_front_rh[-1])],
            "rear_rh_range": [float(shared_rear_rh[0]), float(shared_rear_rh[-1])],
            "grid_shape": list(arrays[f"balance_{wing_angles[0]}"].shape),
        }
        json_path = _AERO_OUT_DIR / f"{car_name}_aero.json"
        json_path.write_text(json.dumps(meta, indent=2) + "\n")

        print(f"  {car_name}: {len(wing_angles)} wing angles → {npz_path.name}")


if __name__ == "__main__":
    parse_all_cars()
